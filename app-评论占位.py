# -*- coding: utf-8 -*-
import streamlit as st
import os
import re
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd
import base64
import requests
import json
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
import tempfile
import shutil
from io import BytesIO

# 页面配置
st.set_page_config(
    page_title="评论情感分析工具",
    page_icon="📊",
    layout="wide"
)

# 负面关键词列表
DEFAULT_NEGATIVE_KEYWORDS = [
    "王总人老实话不多", "人老实话不", "王婆卖瓜", "泰迪", "呼吸",
    "教育消费者", "电是怎么来的", "筷子悬架", "常压油箱", "恒大", "生锈"
]

# API配置 - 从secrets读取，如果没有则使用默认值（本地测试用）
try:
    # 云端部署时使用secrets
    API_KEY = st.secrets["API_KEY"]
    BASE_URL = st.secrets.get("BASE_URL", "https://www.sophnet.com/api/open-apis/v1")
    MODEL = st.secrets.get("MODEL", "Qwen3-VL-235B-A22B-Instruct")
except:
    # 本地测试时使用默认值
    API_KEY = "u6Ow4Rc3S7XU4oSGBxhFjR1sFeSZMf_Jd_UQ2JTdhqx3FhdjSSF8I2vHgrztkru3CXCaWu3Yb_65TmHsqX4RVg"
    BASE_URL = "https://www.sophnet.com/api/open-apis/v1"
    MODEL = "Qwen3-VL-235B-A22B-Instruct"


class WPSImageExtractor:
    """从WPS Excel文件中提取图片"""
    
    def __init__(self, excel_file_path, negative_keywords):
        self.excel_file_path = excel_file_path
        self.file_name = os.path.basename(excel_file_path)
        self.negative_keywords = negative_keywords
    
    def extract_images(self, base_output_dir):
        """提取所有图片并返回图片信息"""
        if not Path(self.excel_file_path).exists():
            return []
        
        file_output_dir = Path(base_output_dir) / self.file_name.replace('.xlsx', '')
        file_output_dir.mkdir(parents=True, exist_ok=True)
        
        try:
            with zipfile.ZipFile(self.excel_file_path, 'r') as zf:
                cellimages_xml_path = 'xl/cellimages.xml'
                if cellimages_xml_path not in zf.namelist():
                    return self._extract_images_alternative(zf, file_output_dir)
                
                rId_to_path = {}
                cellimages_rels_path = 'xl/_rels/cellimages.xml.rels'
                if cellimages_rels_path in zf.namelist():
                    rels_content = zf.read(cellimages_rels_path)
                    rels_root = ET.fromstring(rels_content)
                    for rel in rels_root:
                        rId = rel.get('Id')
                        target_path = rel.get('Target')
                        full_path = str(Path('xl') / Path(target_path)).replace('\\', '/')
                        rId_to_path[rId] = full_path
                
                content = zf.read(cellimages_xml_path)
                root = ET.fromstring(content)
                namespaces = {
                    'etc': 'http://www.wps.cn/officeDocument/2017/etCustomData',
                    'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                    'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
                    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
                }
                
                dispimg_id_to_rid = {}
                for cell_image in root.findall('etc:cellImage', namespaces):
                    cnvpr_tag = cell_image.find('.//xdr:cNvPr', namespaces)
                    blip_tag = cell_image.find('.//a:blip', namespaces)
                    if cnvpr_tag is not None and blip_tag is not None:
                        dispimg_id = cnvpr_tag.get('name')
                        rId = blip_tag.get('{' + namespaces['r'] + '}embed')
                        if dispimg_id and rId:
                            dispimg_id_to_rid[dispimg_id] = rId
                
                final_id_to_path_map = {
                    disp_id: rId_to_path[rId]
                    for disp_id, rId in dispimg_id_to_rid.items() if rId in rId_to_path
                }
                
                if not final_id_to_path_map:
                    return self._extract_images_alternative(zf, file_output_dir)
                
                workbook = openpyxl.load_workbook(self.excel_file_path, data_only=False)
                sheet = workbook.worksheets[0]
                dispimg_pattern = re.compile(r'DISPIMG\(\"([^\"]+)\"')
                images_info = []
                
                for row_index, row in enumerate(sheet.iter_rows(min_row=1), 1):
                    row_output_dir = file_output_dir / f"第_{row_index}_行"
                    images_in_row = []
                    
                    for cell in row:
                        if isinstance(cell.value, str):
                            match = dispimg_pattern.search(cell.value)
                            if match:
                                dispimg_id = match.group(1)
                                if dispimg_id in final_id_to_path_map:
                                    image_path_in_zip = final_id_to_path_map[dispimg_id]
                                    try:
                                        image_data = zf.read(image_path_in_zip)
                                        images_in_row.append((cell.coordinate, image_data, Path(image_path_in_zip).suffix))
                                    except KeyError:
                                        pass
                    
                    if images_in_row:
                        row_output_dir.mkdir(parents=True, exist_ok=True)
                        for coordinate, data, suffix in images_in_row:
                            output_filename = f"{coordinate}{suffix}"
                            output_path = row_output_dir / output_filename
                            with open(output_path, 'wb') as img_f:
                                img_f.write(data)
                            
                            row_data = self._get_row_data(sheet, row_index)
                            images_info.append({
                                'row': row_index,
                                'cell': coordinate,
                                'image_path': str(output_path),
                                'image_data': data,
                                'row_data': row_data,
                                'file_name': self.file_name,
                                'sheet': sheet,
                                'workbook': workbook
                            })
                
                return images_info
                
        except Exception as e:
            st.error(f"提取图片失败: {e}")
            return []
    
    def _extract_images_alternative(self, zip_file, output_dir):
        """备用提取方法"""
        images_info = []
        media_files = [f for f in zip_file.namelist() if f.startswith('xl/media/')]
        
        if not media_files:
            return []
        
        workbook = openpyxl.load_workbook(self.excel_file_path, data_only=False)
        sheet = workbook.worksheets[0]
        
        for i, media_file in enumerate(media_files):
            try:
                image_data = zip_file.read(media_file)
                row_output_dir = output_dir / f"第_{i + 1}_行"
                row_output_dir.mkdir(parents=True, exist_ok=True)
                ext = Path(media_file).suffix
                output_filename = f"图片_{i + 1}{ext}"
                output_path = row_output_dir / output_filename
                
                with open(output_path, 'wb') as img_f:
                    img_f.write(image_data)
                
                row_data = self._get_row_data(sheet, 1)
                images_info.append({
                    'row': 1,
                    'cell': f'A{i + 1}',
                    'image_path': str(output_path),
                    'image_data': image_data,
                    'row_data': row_data,
                    'file_name': self.file_name,
                    'sheet': sheet,
                    'workbook': workbook
                })
            except Exception as e:
                pass
        
        return images_info
    
    def _get_row_data(self, sheet, row_index):
        """获取行数据"""
        row_data = {}
        for col in range(1, sheet.max_column + 1):
            header = sheet.cell(row=1, column=col).value
            value = sheet.cell(row=row_index, column=col).value
            if header:
                row_data[header] = value
        return row_data


class CommentAnalyzer:
    """API分析器"""
    
    def __init__(self, api_key, base_url, model, negative_keywords):
        self.api_key = api_key
        self.base_url = base_url
        self.model = model
        self.negative_keywords = negative_keywords
        self.headers = {
            "Authorization": f"Bearer {self.api_key}",
            "Content-Type": "application/json"
        }
    
    def image_data_to_base64(self, image_data):
        return base64.b64encode(image_data).decode('utf-8')
    
    def analyze_comments(self, image_data, max_retries=3):
        """分析图片中的评论"""
        base64_image = self.image_data_to_base64(image_data)
        
        prompt = f"""请识别这张截图中的所有评论，并分析每条评论的正负面情感。

        注意：
        1. 这是一张比亚迪汽车相关评论的截图
        2. 如果评论包含以下关键词，请直接判断为负面：
           {', '.join(self.negative_keywords)}
        3. 请按顺序列出所有评论，并为每条评论标注：正面/负面
        4. 如果图片中没有评论内容，请返回空列表

        请以JSON格式返回，格式如下：
        {{
            "comments": [
                {{"text": "评论内容1", "sentiment": "正面/负面"}},
                ...
            ]
        }}
        """
        
        payload = {
            "model": self.model,
            "messages": [
                {
                    "role": "user",
                    "content": [
                        {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{base64_image}"}},
                        {"type": "text", "text": prompt}
                    ]
                }
            ],
            "max_tokens": 2000
        }
        
        for attempt in range(max_retries):
            try:
                response = requests.post(
                    f"{self.base_url}/chat/completions",
                    headers=self.headers,
                    json=payload,
                    timeout=30
                )
                
                if response.status_code == 200:
                    result = response.json()
                    content = result['choices'][0]['message']['content']
                    json_match = re.search(r'\{.*\}', content, re.DOTALL)
                    if json_match:
                        return json.loads(json_match.group())
                    else:
                        return {"comments": []}
                else:
                    time.sleep(2)
                    
            except Exception as e:
                time.sleep(2)
        
        return {"comments": []}


class MetricCalculator:
    """指标计算器"""
    
    @staticmethod
    def calculate_positive_ratio(comments, top_n=None):
        if not comments:
            return 0
        if top_n and top_n <= len(comments):
            target_comments = comments[:top_n]
        else:
            target_comments = comments
        positive_count = sum(1 for c in target_comments if c.get('sentiment') == '正面')
        return round((positive_count / len(target_comments) * 100), 2) if target_comments else 0
    
    @staticmethod
    def calculate_standard_metrics(comments):
        standard_metrics = {}
        total = len(comments)
        
        for n, key in [(3, '前3条非负率'), (10, '前10条非负率'), (25, '前25条非负率')]:
            if total >= n:
                ratio = MetricCalculator.calculate_positive_ratio(comments, n)
            else:
                ratio = MetricCalculator.calculate_positive_ratio(comments, total)
            standard_metrics[key] = f"{ratio}%"
        
        return standard_metrics


def process_single_image(image_info, analyzer):
    """处理单张图片"""
    result = analyzer.analyze_comments(image_info['image_data'])
    comments = result.get('comments', [])
    
    calculator = MetricCalculator()
    standard_metrics = calculator.calculate_standard_metrics(comments)
    
    return {
        'file_name': image_info['file_name'],
        'row': image_info['row'],
        'cell': image_info['cell'],
        'comments': comments,
        'standard_metrics': standard_metrics,
        'row_data': image_info['row_data'],
        'image_path': image_info['image_path']
    }


def write_results_to_excel(results, original_file_path):
    """将结果写入Excel"""
    workbook = openpyxl.load_workbook(original_file_path)
    sheet = workbook.worksheets[0]
    
    headers_row = 1
    existing_headers = {}
    for col in range(1, sheet.max_column + 1):
        header = sheet.cell(row=headers_row, column=col).value
        if header:
            existing_headers[header] = col
    
    standard_headers = ['前3条非负率', '前10条非负率', '前25条非负率']
    next_col = sheet.max_column + 1
    header_to_col = dict(existing_headers)
    
    for header in standard_headers:
        if header not in header_to_col:
            sheet.cell(row=headers_row, column=next_col, value=header)
            header_to_col[header] = next_col
            next_col += 1
    
    row_results = {}
    for result in results:
        row_num = result['row']
        if row_num not in row_results:
            row_results[row_num] = {}
        row_results[row_num].update(result['standard_metrics'])
    
    for row_num, metrics in row_results.items():
        for metric_name, value in metrics.items():
            if metric_name in header_to_col:
                sheet.cell(row=row_num, column=header_to_col[metric_name], value=value)
    
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def generate_summary_report(results, original_filename):
    """生成汇总报告"""
    data = []
    for result in results:
        comments = result.get('comments', [])
        positive_count = sum(1 for c in comments if c.get('sentiment') == '正面')
        
        row = {
            '行号': result['row'],
            '单元格': result['cell'],
            '评论总数': len(comments),
            '正面评论数': positive_count,
            '负面评论数': len(comments) - positive_count,
            '整体正面占比': f"{positive_count / len(comments) * 100:.2f}%" if comments else "0%"
        }
        for key, val in result['standard_metrics'].items():
            row[key] = val
        data.append(row)
    
    df = pd.DataFrame(data)
    output = BytesIO()
    df.to_excel(output, index=False)
    output.seek(0)
    return output


def generate_details_report(results, original_filename):
    """生成评论详情报告"""
    details_data = []
    for result in results:
        comments = result.get('comments', [])
        for i, comment in enumerate(comments, 1):
            details_data.append({
                '行号': result['row'],
                '单元格': result['cell'],
                '评论序号': i,
                '评论内容': comment.get('text', ''),
                '情感': comment.get('sentiment', '')
            })
    
    df = pd.DataFrame(details_data) if details_data else pd.DataFrame(columns=['行号', '单元格', '评论序号', '评论内容', '情感'])
    output = BytesIO()
    df.to_excel(output, index=False)
    output.seek(0)
    return output


# ==================== Streamlit UI ====================

st.title("📊 评论情感分析工具")
st.markdown("从WPS Excel文件中提取截图，使用AI分析评论情感，自动计算正面占比")

# 初始化session state
if 'negative_keywords' not in st.session_state:
    st.session_state.negative_keywords = DEFAULT_NEGATIVE_KEYWORDS.copy()
if 'results_data' not in st.session_state:
    st.session_state.results_data = None
if 'processed' not in st.session_state:
    st.session_state.processed = False

with st.sidebar:
    st.header("⚙️ 配置")
    
    # 显示API状态（不显示密钥本身）
    try:
        if st.secrets.get("API_KEY"):
            st.success("✅ API密钥已配置")
    except:
        st.warning("⚠️ 未配置API密钥，使用默认值（仅限测试）")
    
    max_workers = st.slider(
        "并发处理数",
        min_value=1,
        max_value=5,  # 云端限制低一些，避免超时
        value=3,
        help="同时处理的图片数量"
    )
    
    st.markdown("---")
    st.markdown("### 负面关键词")
    keywords_text = st.text_area(
        "编辑负面关键词（每行一个）",
        value="\n".join(st.session_state.negative_keywords),
        height=200
    )
    
    if st.button("更新关键词"):
        st.session_state.negative_keywords = [k.strip() for k in keywords_text.split("\n") if k.strip()]
        st.success(f"已更新 {len(st.session_state.negative_keywords)} 个关键词")

# 主界面
uploaded_file = st.file_uploader(
    "上传Excel文件（支持WPS格式，内含DISPIMG图片）",
    type=['xlsx'],
    help="上传包含截图评论的Excel文件"
)

if uploaded_file is not None:
    st.session_state.processed = False
    st.session_state.results_data = None
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
        tmp_file.write(uploaded_file.getvalue())
        temp_file_path = tmp_file.name
    
    st.success(f"文件已上传: {uploaded_file.name}")
    
    col1, col2 = st.columns(2)
    with col1:
        st.metric("文件大小", f"{len(uploaded_file.getvalue()) / 1024:.1f} KB")
    
    if st.button("🚀 开始分析", type="primary"):
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        try:
            status_text.text("正在提取图片...")
            extractor = WPSImageExtractor(temp_file_path, st.session_state.negative_keywords)
            temp_dir = tempfile.mkdtemp()
            images = extractor.extract_images(temp_dir)
            
            if not images:
                st.error("未找到任何图片，请检查Excel文件格式")
                st.stop()
            
            st.info(f"找到 {len(images)} 张图片")
            progress_bar.progress(10)
            
            analyzer = CommentAnalyzer(API_KEY, BASE_URL, MODEL, st.session_state.negative_keywords)
            
            status_text.text("正在分析图片（这可能需要几分钟）...")
            results = []
            processed = 0
            
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                futures = {executor.submit(process_single_image, img, analyzer): img for img in images}
                
                for future in as_completed(futures):
                    result = future.result()
                    if result:
                        results.append(result)
                    processed += 1
                    progress = 10 + int(80 * processed / len(images))
                    progress_bar.progress(progress)
                    status_text.text(f"已分析 {processed}/{len(images)} 张图片")
            
            if not results:
                st.error("没有成功分析任何图片")
                st.stop()
            
            progress_bar.progress(95)
            status_text.text("正在生成报告...")
            
            output_excel = write_results_to_excel(results, temp_file_path)
            summary_report = generate_summary_report(results, uploaded_file.name)
            details_report = generate_details_report(results, uploaded_file.name)
            
            st.session_state.results_data = {
                'output_excel': output_excel,
                'summary_report': summary_report,
                'details_report': details_report,
                'file_name': uploaded_file.name,
                'total_images': len(results),
                'total_comments': sum(len(r.get('comments', [])) for r in results)
            }
            st.session_state.processed = True
            
            progress_bar.progress(100)
            status_text.text("分析完成！")
            
        except Exception as e:
            st.error(f"处理失败: {str(e)}")
            import traceback
            st.code(traceback.format_exc())
        finally:
            if os.path.exists(temp_file_path):
                os.unlink(temp_file_path)
    
    # 显示结果和下载按钮
    if st.session_state.processed and st.session_state.results_data:
        data = st.session_state.results_data
        
        st.success(f"✅ 分析完成！共处理 {data['total_images']} 张图片，识别到 {data['total_comments']} 条评论")
        
        st.markdown("### 📥 下载结果文件")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            output_excel = data['output_excel']
            output_excel.seek(0)
            b64_original = base64.b64encode(output_excel.getvalue()).decode()
            download_link_original = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64_original}" download="{data["file_name"].replace(".xlsx", "")}_已分析.xlsx" style="display: inline-block; background-color: #ff4b4b; color: white; padding: 10px 20px; text-align: center; text-decoration: none; border-radius: 5px; width: 100%;">📊 原始文件（已写入结果）</a>'
            st.markdown(download_link_original, unsafe_allow_html=True)
            st.caption("包含原始数据和计算结果")
        
        with col2:
            summary_report = data['summary_report']
            summary_report.seek(0)
            b64_summary = base64.b64encode(summary_report.getvalue()).decode()
            download_link_summary = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64_summary}" download="{data["file_name"].replace(".xlsx", "")}_汇总报告.xlsx" style="display: inline-block; background-color: #ff4b4b; color: white; padding: 10px 20px; text-align: center; text-decoration: none; border-radius: 5px; width: 100%;">📈 汇总报告</a>'
            st.markdown(download_link_summary, unsafe_allow_html=True)
            st.caption("每行图片的统计汇总")
        
        with col3:
            details_report = data['details_report']
            details_report.seek(0)
            b64_details = base64.b64encode(details_report.getvalue()).decode()
            download_link_details = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64_details}" download="{data["file_name"].replace(".xlsx", "")}_评论详情.xlsx" style="display: inline-block; background-color: #ff4b4b; color: white; padding: 10px 20px; text-align: center; text-decoration: none; border-radius: 5px; width: 100%;">💬 评论详情</a>'
            st.markdown(download_link_details, unsafe_allow_html=True)
            st.caption("每条评论的详细内容")
        
        st.markdown("---")
        st.info("💡 提示：点击按钮即可下载，不会跳转页面，可以依次下载所有文件")

else:
    st.info("👈 请上传Excel文件开始分析")
    
    st.markdown("""
    ### 使用说明
    
    1. **上传文件**：点击上方上传按钮，选择包含DISPIMG图片的WPS Excel文件
    2. **开始分析**：点击"开始分析"按钮
    3. **下载结果**：分析完成后，点击三个下载按钮
    
    ### 输出文件说明
    
    | 文件名 | 内容说明 |
    |--------|---------|
    | *_已分析.xlsx | 原文件 + 计算结果（新增三列） |
    | *_汇总报告.xlsx | 每行图片的统计汇总 |
    | *_评论详情.xlsx | 每条评论的详细内容 |
    
    ### 注意事项
    
    - 处理时间取决于图片数量，请耐心等待
    - 建议单次处理不超过50张图片
    - 如果处理失败，请刷新页面重试
    """)

st.markdown("---")
st.markdown("💡 提示：点击下载按钮不会跳转页面，可以连续下载多个文件")